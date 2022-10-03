from bs4 import BeautifulSoup
import requests
import openpyxl
from openpyxl.styles import Font


excel = openpyxl.Workbook()


def topics_choosing():
	'''Allowes user to select a topic they are interested in.
		Checks, wheather the input value is valid or not.'''
	global news_topics
	news_topics = ['My country news', 'World', 'My local news', 'Business', 'Technologies',
					   'Entertainment', 'Sports', 'Science', 'Health', 'Everything']
	topic = ''

	print('Please, specify the topic you are interested in.')
	print('------------------------------------------')
	print('Available topics:')

	for i in range(len(news_topics)):
		print(f'{i+1}. {news_topics[i]}')

	print('')

	while True:
		topic = input(': ')

		if topic not in news_topics:
			print('Please, specify the right topic.\n')
		else:
			print('----------------------------------')
			return topic


def url_setting(topic):
	'''Sets URLs of available topics.'''

	url = 'https://news.google.com/topics/.'
	r = requests.get(url)
	soup =  BeautifulSoup(r.text, 'lxml')

	topic_list = soup.find('div', jsname='V2bVMb')
	path_tags = topic_list.find_all('a', class_='SFllF')
	path = []

	for tag in range(1, len(path_tags)):
		path.append(path_tags[tag]['href'])

	final_url = {
		'My country news': '',
		'World': '',
		'My local news': '',
		'Business': '',
		'Technologies': '',
		'Entertainment': '',
		'Sports': '',
		'Science': '',
		'Health': ''
		}

	final_url_list = list(final_url)
	
	for i in range(len(final_url)):
		final_url.update({final_url_list[i] : url + path[i]})

	return final_url.get(topic)


def news_scraping(topic_url, sheet):
	'''Takes a topic as an input.
		Scrapes title, news company, publishing time and link.'''

	r = requests.get(topic_url)
	soup = BeautifulSoup(r.text, 'lxml')

	headlines = soup.find_all('article', class_='MQsxIb xTewfe R7GTQ keNKEd j7vNaf Cc0Z5d EjqUne')

	for news in headlines:
		title = news.h3.a.text
		news_company = news.div.div.a.text
		publishing_time = news.div.div.time.text
		link = 'https://news.google.com/topics/.' + news.a['href']

		sheet.append([title, news_company, publishing_time, link])


def data_to_excel_everything():
	'''Creates an excel file with every listed topic.'''

	ft = Font(bold=True)

	for i in range(9):
		sheet = excel.create_sheet(title=news_topics[i])
		sheet.append(['Title', 'News Company', 'Publication Time', 'Link'])
		news_scraping(url_setting(news_topics[i]), sheet)

		sheet['A1'].font = ft
		sheet['B1'].font = ft
		sheet['C1'].font = ft
		sheet['D1'].font = ft

		sheet.column_dimensions['A'].width = 150
		sheet.column_dimensions['B'].width = 25
		sheet.column_dimensions['C'].width = 20
		sheet.column_dimensions['D'].width = 255

		print(f'"{news_topics[i]}" topic was succesfully added.')

	del excel['Sheet']

	excel.save('The Google News (Everything).xlsx')

	print('Execution done.')


def data_to_excel_particular(topic):
	'''Creates an excel file with a particular topic.'''

	ft = Font(bold=True)

	sheet = excel.create_sheet(title=topic)
	sheet.append(['Title', 'News Company', 'Publication Time', 'Link'])
	news_scraping(url_setting(topic), sheet)

	sheet['A1'].font = ft
	sheet['B1'].font = ft
	sheet['C1'].font = ft
	sheet['D1'].font = ft

	del excel['Sheet']

	sheet.column_dimensions['A'].width = 150
	sheet.column_dimensions['B'].width = 25
	sheet.column_dimensions['C'].width = 20
	sheet.column_dimensions['D'].width = 255

	excel.save(f'The Google News ({topic}).xlsx')

	print(f'"{topic}" topic was succesfully added.')
	print('Execution done.')


topic = topics_choosing()

if topic == 'Everything':
	data_to_excel_everything()
	input('\nPress ENTER to exit')
else:
	data_to_excel_particular(topic)
	input('\nPress ENTER to exit')